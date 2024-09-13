import openpyxl
from openpyxl import load_workbook
import pyautogui
import time


pyautogui.PAUSE = 0.3

planilha_contabil = load_workbook('RAZAO_CONTABIL.xlsx', data_only=True)  # Use data_only=True para obter os valores calculados
contabil_sheet = planilha_contabil['contabil']

for linha in contabil_sheet.iter_rows(min_row=2):
    data_inicio = contabil_sheet.cell(row=linha[0].row, column=1).value  # Coluna A
    data_final = contabil_sheet.cell(row=linha[0].row, column=2).value  # Coluna B
    eg_inicio = contabil_sheet.cell(row=linha[0].row, column=3).value  # Coluna C
    eg_final = contabil_sheet.cell(row=linha[0].row, column=4).value  # Coluna D
    nome_arquivo = contabil_sheet.cell(row=linha[0].row, column=5).value  # Coluna E

    print(data_inicio)
    print(data_final)
    print(eg_inicio)
    print(eg_final)
    print(nome_arquivo)
